import pandas as pd
from zipfile import BadZipFile
## Removed unused import os
from django.conf import settings
from pathlib import Path

XLSX = Path(settings.BASE_DIR) / "data" / "cafelytics_data_ML.xlsx"
if not XLSX.exists():
    raise FileNotFoundError(f"Excel file not found: {XLSX}")

def load_menu_df():
    try:
        df = pd.read_excel(XLSX, sheet_name='Data', engine='openpyxl')
    except BadZipFile:
        print(f"Error: The Excel file '{XLSX}' is not a valid zip file or is corrupted.")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error reading Excel file '{XLSX}': {e}")
        return pd.DataFrame()
    # Normalize columns
    df.columns = [c.strip() for c in df.columns]
    col_map = {
        'Item ID': 'ItemID',
        'Item Name': 'ItemName',
        'Preference Score': 'PreferenceScore',
        'Price': 'Price',
        'Type': 'Type',
        'Category': 'Category',
        'Availability': 'Availability'
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
    for col in ['ItemID', 'ItemName', 'Price', 'PreferenceScore', 'Type', 'Category', 'Availability']:
        if col not in df.columns:
            df[col] = None
    df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0)
    df['PreferenceScore'] = pd.to_numeric(df['PreferenceScore'], errors='coerce').fillna(0)
    df['Type'] = df['Type'].astype(str).fillna('Unknown')
    df['Category'] = df['Category'].astype(str).fillna('General')
    df['Availability'] = df['Availability'].astype(str).fillna('Yes')
    df['ItemID'] = df['ItemID'].astype(str).fillna('')
    df['ItemName'] = df['ItemName'].astype(str).fillna('')
    return df
    if 'PreferenceScore' not in df.columns:
        df['PreferenceScore'] = 0.0
    df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0.0)
    df['Type'] = df['Type'].astype(str).str.title().fillna('Veg')
    df['Category'] = df['Category'].astype(str).fillna('Other')
    df['Availability'] = df['Availability'].astype(str).fillna('All Day')
    return df


def write_menu_df(df):
    # caution: overwrites sheet 'Data'
    from openpyxl import load_workbook
    book = load_workbook(XLSX)
    with pd.ExcelWriter(XLSX, engine='openpyxl') as writer:
        writer.book = book
        if 'Data' in book.sheetnames:
            idx = book.sheetnames.index('Data')
            book.remove(book.worksheets[idx])
        df.to_excel(writer, sheet_name='Data', index=False)
        writer.save()
